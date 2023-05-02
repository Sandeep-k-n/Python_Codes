#!/usr/bin/env python
# coding: utf-8


import time
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, StaleElementReferenceException
from selenium.webdriver.support.expected_conditions import staleness_of
import pandas as pd
import numpy as np
from datetime import date, datetime
from selenium.webdriver.chrome.options import Options
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# option = Options()
# option.add_argument("--disable-popup-blocking")
#driver = webdriver.Chrome("yourpathtochromedriver",chrome_options=option)

#start time
start = time.time()

# Initializing Chrome reference
driver = webdriver.Chrome("C:\\chromedriver\\chromedriver.exe")

#Specify Search URL
driver.get("https://pieweb.plasteurope.com/default.aspx?pageid=22223")
driver.implicitly_wait(30)
driver.switch_to.frame(driver.find_element_by_tag_name("iframe"))
driver.find_element_by_xpath('//*[@id="accept_all"]').click()

# Defining Username Password
Username = 'p.mamta.rao@consultant.volvo.com'
Password = 'Mamt@1802' 

driver.implicitly_wait(30)
element_to_hover_over = driver.find_element(By.XPATH,"/html/body/form/span[1]/div/div[1]/div[2]/div/ul/li[1]/a")
hover = ActionChains(driver).move_to_element(element_to_hover_over)
hover.perform()

# find username/email field and send the username itself to the input field
driver.find_element(By.NAME,"ctl04$Header$ctl01$textbox_username").send_keys(Username)
# find password input field and insert password as well
driver.find_element(By.NAME,"ctl04$Header$ctl01$textbox_password").send_keys(Password)
driver.find_element(By.NAME,"ctl04$Header$ctl01$checkbox_permanent").click()
driver.find_element(By.NAME,"ctl04$Header$ctl01$button_login").click()

driver.implicitly_wait(30)

# Get links and names of materials
Material_List = [element.get_attribute("text") for element in driver.find_elements(By.XPATH,'//div[@id="layout_wrapper"]/div[@id="layout_spalte_100"]/table[@class="pp-detail-liste"]/tbody/tr/td/a')]
driver.refresh()
links = [my_elem.get_attribute("href") for my_elem in driver.find_elements(By.XPATH,'//div[@id="layout_wrapper"]/div[@id="layout_spalte_100"]/table[@class="pp-detail-liste"]/tbody/tr/td/a[1]')]
driver.get(links[0])

Value_List = [my_elem.get_attribute("value") for my_elem in driver.find_elements(By.XPATH,'/html/body/form/div[4]/div/div/p[1]/select/option')]
Value_text_List = [my_elem.get_attribute("text") for my_elem in driver.find_elements(By.XPATH,'/html/body/form/div[4]/div/div/p[1]/select/option')] 
lnk = "https://pieweb.plasteurope.com/default.aspx?pageid=22222&typ="
df_Values = pd.DataFrame(list(zip(Value_List, Value_text_List)), columns =['Value', 'Material'])
df_Values = df_Values.query('Value.astype("str").str.len().astype("str") > "2"')
df_Values['link'] = lnk + df_Values['Value'].astype(str)
df_Values.reset_index(inplace = True)


Polymer_Data = pd.DataFrame(columns=['DATE'])
for i in range(len(df_Values)) :
  linkcolumn = df_Values.loc[i, "link"]
  linkMaterial = df_Values.loc[i, "Material"]
  link = linkcolumn + '#tab2monate-tab'
  driver.get(link)
  
  page_source = pd.read_html(driver.page_source)
  for dfs in page_source:
      if len(dfs) > 45:
          df = dfs
      else:
          df = pd.DataFrame(columns=['DATE', "Material"])

  df = df.iloc[:,[0,3]]
  df.columns = ['DATE', linkMaterial]
  df = df.drop([0,1, (len(df)-1)])
  Polymer_Data = df.merge(Polymer_Data, on='DATE', how='outer')

Polymer_Data["Date"] = Polymer_Data["DATE"].apply(pd.to_datetime)
# Polymer_Data.sort_values(by="Date",ascending=True)


Pie_web = Polymer_Data.sort_values(by="Date",ascending=True).set_index("Date").last("13M")
Pie_web["Date"] = Pie_web["DATE"].apply(pd.to_datetime)
Pie_web["Period"] = Pie_web['Date'].dt.strftime('%Y%m')
Pie_web = Pie_web.set_index(['Period'])
Pie_web.drop(Pie_web.columns[[0,120]], axis=1, inplace=True)
Pie_web.dropna(how='all', axis=1, inplace=True)
Index_last = []
for columns in Pie_web:
    Index_last.append(Pie_web[columns].last_valid_index())

#test for new format
def prepend_index_level(index, key, name=None):
    names = index.names
    if index.nlevels==1:
        # Sequence of tuples
        index = ((item,) for item in index)

    tuples_gen = ((key,)+item for item in index)
    return pd.MultiIndex.from_tuples(tuples_gen, names=[name]+names)

# Add Multi Index
Previous_month = datetime.now().replace(month=int(datetime.now().strftime('%m'))).strftime('%Y%m')
Pie_web.columns = prepend_index_level(Pie_web.columns, key= Previous_month, name="LastActualPeriod")
Pie_web.columns = prepend_index_level(Pie_web.columns, key="ton", name="UOMCode")
Pie_web.columns = prepend_index_level(Pie_web.columns, key="EUR", name="CurrencyCode")
Pie_web.columns = prepend_index_level(Pie_web.columns, key="PIEWEB", name="Index Type")
Pie_web.columns = prepend_index_level(Pie_web.columns, key="N", name="Standard")

Pie_web.columns = Pie_web.columns.rename("RMCode", level=5)

# Reorder Multi Index
Pie_web = Pie_web.reorder_levels(["Standard", "Index Type", "RMCode", "CurrencyCode", "UOMCode", "LastActualPeriod"], axis=1)

# Multi Index to rows( To save as excel - Excel doesn't support multi index)
Pie_web = Pie_web.columns.to_frame().T.append(Pie_web, ignore_index=False)
Pie_web.columns = range(len(Pie_web.columns))
Pie_web.iloc[5] = Index_last

# Load to Openpyxl workbook
wb = Workbook()
ws = wb.active

# Load Dataframe to Workbook
for r in dataframe_to_rows(Pie_web, index=True, header=True):
    ws.append(r)

# Color the cells based on condition     
bd = Side(style='thin', color="000000")
for rows in ws.iter_rows():
   for cell in rows:
     cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
     cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
     if cell.row >= 3  and cell.row <= 8:
         cell.fill = PatternFill(start_color="00C0C0C0", end_color ="00C0C0C0", fill_type = "solid")
         cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
     elif cell.row >= 8 and cell.row <= (len(Pie_web) + 2):
         cell.fill = PatternFill(start_color="00CCFFCC", end_color ="00CCFFCC", fill_type = "solid")
     
# Save workbook to Excel
wb.save("PIEWEB.xlsx")
# Pie_web = Pie_web.melt(id_vars=["DATE"], 
#         var_name="Raw_Material", 
#         value_name="Amount")

# Pie_web["Date"] = Pie_web["DATE"].apply(pd.to_datetime)
# Pie_web["Forecast_Actual"] = np.where(Pie_web.Date < datetime.strptime(datetime.today().strftime("%Y-%m-01"), '%Y-%m-%d'), 'A', 'F')
# Pie_web["Version"] = datetime.now().strftime('%Y%m')
# Pie_web["Period"] = Pie_web['Date'].dt.strftime('%Y%m')

# Pie_web = Pie_web.assign(Currency= "EUR")   
# Pie_web = Pie_web.assign(Unit_Of_Weight= "TON") 
# Pie_web = Pie_web.assign(Index = "PIE WEB") 
# Pie_web = Pie_web.dropna(subset=['Amount'])

# Pie_web = Pie_web[["Index", "Raw_Material", "Currency", "Unit_Of_Weight", "Period", "Version", "Amount", "Forecast_Actual"]]


# del Polymer_Data['Date']

# Polymer_Data = Polymer_Data.melt(id_vars=["DATE"], 
#         var_name="Raw_Material", 
#         value_name="Amount")

# Recyclate_Material = Recyclate_Material.merge(Polymer_Data, on='Raw_Material', how='inner')
# Virgin_Material = (Recyclate_Material.merge(Polymer_Data, on='Raw_Material', how='right',indicator=True)\
#                     .query('_merge == "right_only"')\
#                     .drop('_merge', 1))

# Virgin_Material.rename(columns = {'DATE_y':'DATE', 'Amount_y':'Amount'}, inplace = True)

# Virgin_Material = Virgin_Material[['Raw_Material','DATE','Amount']]

# # determining the name of the file
# file_name = 'PIE_WEB.xlsx'
# recyclate_name = 'Recyclate_Material.xlsx'
# virgin_mat_name = 'Virgin_Material.xlsx'
      
# # saving the excel
# Pie_web.to_excel(file_name, index=False)
# Recyclate_Material.to_excel(recyclate_name, index=False)
# Virgin_Material.to_excel(virgin_mat_name , index=False)

# #end time
# end = time.time()

# # total time taken
# print(f"Total runtime of the program is {end - start}")

