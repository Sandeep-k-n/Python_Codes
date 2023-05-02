# -*- coding: utf-8 -*-
"""
Created on Fri Dec  3 18:47:32 2021

@author: a383785
"""
import pandas as pd
#import numpy as np
import selenium
from selenium import webdriver
import time
import datetime
from datetime import date, datetime
#import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
# from selenium.webdriver.common.keys import Keys

#Install Driver
driver = webdriver.Chrome("C:\\chromedriver\\chromedriver.exe")

#start time
start = time.time()

# credentials
username = "ramesh.s@volvo.com"
password = "Classic123#"

driver.get("https://portal.mepsinternational.com/?logout=true")

# find username/email field and send the username itself to the input field
driver.find_element_by_id("Reg_Email").send_keys(username)
# find password input field and insert password as well
driver.find_element_by_id("Reg_Password").send_keys(password)
# click login button
driver.find_element_by_xpath("//*[@type='submit']").click()
time.sleep(2)
# click view subscriptions
driver.find_element_by_xpath('//a[@href="/account/subscriptions/"]').click()

list_of_hrefs = []

content_blocks = driver.find_elements_by_class_name("table-primary__body")

for block in content_blocks:
    elements = block.find_elements_by_tag_name("a")
    for el in elements:
        test = el.get_attribute("href")
        if (test != "https://portal.mepsinternational.com/account/single-licence/" ) and (test not in list_of_hrefs):
            list_of_hrefs.append(el.get_attribute("href"))
            
df = pd.DataFrame(columns = ["DATE"])
        
for lst in list_of_hrefs:
    driver.get(lst)
    time.sleep(2)
    df1 = pd.DataFrame() 
    
    Title = []     
    dates = []
    values = []
    # to identify the table column
    l=driver.find_elements_by_xpath ("//*[@class= 'table-primary']/tbody/tr/td[2]")
    # to traverse through the list of cell data
    for i in l :
        # to get the cell data with text method
        values.append(i.text)
        
    l=driver.find_elements_by_xpath ("//*[@class= 'table-primary']/tbody/tr/td[1]")
    # to traverse through the list of cell data
    for i in l :
        # to get the cell data with text method
        dates.append(i.text)    

    # to identify the table column headers in row 1
    header_l=driver.find_elements_by_xpath ("//*[@class= 'table-primary']/thead[1]/tr[1]/th")
    # to traverse through the list of headers
    for i in header_l :
        # to get the header with text method
        Title.append(i.text)
        
    # both lists, with columns specified
    df1 = pd.DataFrame(list(zip(dates, values)), columns =[Title[0], Title[1]])
    
    df = df.merge(df1, on='DATE', how='outer')

df['Date'] = pd.to_datetime(df['DATE'], errors='coerce')
df["Period"] = df['Date'].dt.strftime('%Y%m')
df = df.sort_values(by="Date",ascending=True).set_index("Date").last("24M")
df = df.set_index(['Period'])
# Remove column name 'Date', 'DATE'
df.drop(df.columns[[0]], axis=1, inplace=True)
#test for new format
def prepend_index_level(index, key, name=None):
    names = index.names
    if index.nlevels==1:
        # Sequence of tuples
        index = ((item,) for item in index)

    tuples_gen = ((key,)+item for item in index)
    return pd.MultiIndex.from_tuples(tuples_gen, names=[name]+names)

# Add Multi Index
Previous_month = datetime.now().replace(month=int(datetime.now().strftime('%m')) - 1).strftime('%Y%m')
df.columns = prepend_index_level(df.columns, key= Previous_month, name="LastActualPeriod")
df.columns = prepend_index_level(df.columns, key="ton", name="UOMCode")
df.columns = prepend_index_level(df.columns, key="EUR", name="CurrencyCode")
df.columns = prepend_index_level(df.columns, key="MEPS", name="Index Type")
df.columns = prepend_index_level(df.columns, key="N", name="Standard")

df.columns = df.columns.rename("RMCode", level=5)

# Reorder Multi Index
df = df.reorder_levels(["Standard", "Index Type", "RMCode", "CurrencyCode", "UOMCode", "LastActualPeriod"], axis=1)

# Multi Index to rows( To save as excel - Excel doesn't support multi index)
df = df.columns.to_frame().T.append(df, ignore_index=False)
df.columns = range(len(df.columns))

# Load to Openpyxl workbook
wb = Workbook()
ws = wb.active

# Load Dataframe to Workbook
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

# Find previous month row
col = []
for rows in ws.iter_rows( max_col=1):
   for cell in rows:         
        if cell.value == Previous_month: 
            col.append(cell.row)

# Color the cells based on condition     
bd = Side(style='thin', color="000000")
for rows in ws.iter_rows():
   for cell in rows:
     cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
     cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
     if cell.row >= 3  and cell.row <= 8:
         cell.fill = PatternFill(start_color="00C0C0C0", end_color ="00C0C0C0", fill_type = "solid")
         cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
     elif cell.row >= 8 and cell.row <= int(col[0]):
         cell.fill = PatternFill(start_color="00CCFFCC", end_color ="00CCFFCC", fill_type = "solid")
     elif cell.row > int(col[0]) and cell.row <= (len(df) + 2):
         cell.fill = PatternFill(start_color="00FFCC99", end_color ="00FFCC99", fill_type = "solid")

# Save workbook to Excel
wb.save("MEPS.xlsx")
      
# col = ws.column_dimensions['A']
# col.font = Font(bold=True)

# row = ws.row_dimensions[9]
# row.font = Font(underline="single")
# bd = Side(style='thin', color="000000")
# row.border = Border(left=bd, top=bd, right=bd, bottom=bd)


# df.to_csv('MEPS.csv')
# df.index.Period = [None]

# df2 = df.copy()
# df = df2.copy()
#df.reset_index()

# # df = df.melt(id_vars=["DATE"], 
# #         var_name="Raw_Material", 
# #         value_name="Amount")
# # converting datatypes
# df = df.astype("int64", errors='ignore')

# # Add column for Actual/Forecast
# today = datetime.now()
# first_day_of_month = today.strftime("%Y-%m-01")
# # df["Date"] = df["DATE"].apply(pd.to_datetime)
# # df = df_test
# df['Date'] = pd.to_datetime(df['DATE'], errors='coerce')
# df["Forecast_Actual"] = np.where(df.Date < datetime.strptime(first_day_of_month, '%Y-%m-%d'), 'A', 'F')
# df["Version"] = today.strftime('%Y%m')
# df["Period"] = df['Date'].dt.strftime('%Y%m')

# df = df.sort_values(by="Date",ascending=True).set_index("Date").last("24M")

# weight_currency = []
# # to identify the weight and currency
# weight_currency_l=driver.find_elements_by_xpath ("//*[@class='data-bar']/p/span  ")
# # to traverse through the list of headers
# for i in weight_currency_l :
#     # to get the header with text method
#     weight_currency.append(i.text)

# # Add Currency, Weight and Index columns
# currency_weight = weight_currency[0].split()
# # df = df.assign(Currency= currency_weight[0])   
# # df = df.assign(Unit_Of_Weight= currency_weight[-1]) 
# df = df.assign(Currency= "EUR")   
# df = df.assign(Unit_Of_Weight= "TON") 
# df = df.assign(Index = "MEPS") 

# df = df[["Index", "Raw_Material", "Currency", "Unit_Of_Weight", "Period", "Version", "Amount", "Forecast_Actual"]]

# df2 = df.copy()

# # determining the name of the file
# file_name = 'MEPS.xlsx'
      
# # saving the excel
# df.to_excel(file_name, index=True)

# #end time
# end = time.time()

# # total time taken
# print(f"Total runtime of the program is {end - start}")

# df.to_csv('MEPS.csv')

# for rows in ws.iter_rows(min_row=3, max_row=8, min_col=1):
#    for cell in rows:
#      cell.fill = PatternFill(start_color="00C0C0C0", end_color ="00C0C0C0", fill_type = "solid")
#      bd = Side(style='thin', color="000000")
#      cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
#      cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
#      # cell.fill = PatternFill(fgColor="	00FFFF00")

            
# for rows in ws.iter_rows(min_row=9, max_row=(len(df) + 2), min_col=1):
#    for cell in rows:
#      cell.fill = PatternFill(start_color="00CCFFCC", end_color ="00CCFFCC", fill_type = "solid")
#      bd = Side(style='thin', color="000000")
#      cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
#      if cell.value == Previous_month: 
#          print(cell.row)
